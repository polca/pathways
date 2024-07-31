import os
import re
from pathlib import Path
from typing import Dict, Set, Tuple
from zipfile import BadZipFile

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook, load_workbook
from SALib.analyze import delta

from pathways.filesystem_constants import STATS_DIR


def log_double_accounting(
    filtered_names: Dict[Tuple[str, ...], Set[str]],
    exception_names: Dict[Tuple[str, ...], Set[str]],
    export_path: Path,
):
    """
    Log the unique names of the filtered activities and exceptions to an Excel file,
    distinguished by categories.

    :param filtered_names: Dictionary of category paths to sets of filtered activity names.
    :param exception_names: Dictionary of category paths to sets of exception names.
    :param export_path: Path to the export Excel file.
    """

    data_filtered = {
        "/".join(category): list(names)
        for category, names in filtered_names.items()
        if names
    }
    data_exceptions = {
        "/".join(category): list(names)
        for category, names in exception_names.items()
        if names
    }

    filtered_df = pd.DataFrame(
        dict([(k, pd.Series(v)) for k, v in data_filtered.items()])
    )
    exception_df = pd.DataFrame(
        dict([(k, pd.Series(v)) for k, v in data_exceptions.items()])
    )

    export_path.parent.mkdir(parents=True, exist_ok=True)

    if export_path.exists():
        try:
            # Load the existing workbook
            with pd.ExcelWriter(
                export_path, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            ) as writer:

                # Remove the existing sheets if they exist
                if "Double accounting - Zeroed" in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index("Double accounting - Zeroed")
                    std = writer.book.worksheets[idx]
                    writer.book.remove(std)
                    writer.book.create_sheet("Double accounting - Zeroed", idx)
                if "Double accounting - Exceptions" in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index("Double accounting - Exceptions")
                    std = writer.book.worksheets[idx]
                    writer.book.remove(std)
                    writer.book.create_sheet("Double accounting - Exceptions", idx)

                # Write DataFrames to the appropriate sheets
                filtered_df.to_excel(
                    writer,
                    sheet_name="Double accounting - Zeroed",
                    index=False,
                )
                exception_df.to_excel(
                    writer,
                    sheet_name="Double accounting - Exceptions",
                    index=False,
                )

        except BadZipFile:
            print(
                f"Warning: '{export_path}' is not a valid Excel file. Creating a new file."
            )
            with pd.ExcelWriter(export_path, engine="openpyxl", mode="w") as writer:
                filtered_df.to_excel(
                    writer, sheet_name="Double accounting - Zeroed", index=False
                )
                exception_df.to_excel(
                    writer, sheet_name="Double accounting - Exceptions", index=False
                )
    else:
        with pd.ExcelWriter(export_path, engine="openpyxl", mode="w") as writer:
            filtered_df.to_excel(
                writer, sheet_name="Double accounting - Zeroed", index=False
            )
            exception_df.to_excel(
                writer, sheet_name="Double accounting - Exceptions", index=False
            )


def log_subshares(
    shares: dict,
    region: str,
) -> pd.DataFrame:
    """
    Create a pandas DataFrame where the keys of shares are the columns
    and the values are the values. The region is added as a column.
    """

    df = pd.DataFrame(shares)
    df["region"] = region
    df["iteration"] = range(1, len(df) + 1)

    return df[["iteration", "region"] + list(shares.keys())]


def log_uncertainty_values(
    region: str,
    uncertainty_indices: np.array,
    uncertainty_values: np.array,
) -> pd.DataFrame:
    """
    Create a pandas DataFrame with the region and uncertainty indices as columns,
    the uncertainty values as values, and the iteration number as the index.

    :param region: Name of the region
    :param uncertainty_indices: Indices of the uncertainty values
    :param uncertainty_values: Uncertainty values
    :return: DataFrame with region as column, uncertainty indices as indices and uncertainty values as values

    """

    # convert 2D numpy array to list of tuples
    uncertainty_indices = uncertainty_indices.tolist()
    uncertainty_indices = [[str(x) for x in index] for index in uncertainty_indices]
    uncertainty_indices = ["::".join(index) for index in uncertainty_indices]

    df = pd.DataFrame(uncertainty_values.T, columns=uncertainty_indices)
    df["region"] = region
    df["iteration"] = range(1, len(df) + 1)

    return df[["iteration", "region"] + uncertainty_indices]


def log_results(
    total_impacts: np.array,
    methods: list,
    region: str,
):
    """
    Log the characterized inventory results for each LCIA method into separate columns in an Excel file.

    :param total_impacts: numpy array of total impacts for each method.
    :param methods: List of method names.
    :param region: Region name as a string.
    """

    df = pd.DataFrame(total_impacts.T, columns=methods)
    df["region"] = region
    df["iteration"] = range(1, len(df) + 1)

    return df[["iteration", "region"] + methods]


def create_mapping_sheet(indices: dict) -> pd.DataFrame:
    """
    Create a mapping sheet for the activities with uncertainties.
    """

    # Converting the dictionary into a pandas DataFrame
    df = pd.DataFrame(indices.items(), columns=["Index", "Value"])

    # Split the 'Index' column into four separate columns
    df[["Name", "Product", "Unit", "Region"]] = pd.DataFrame(
        df["Index"].tolist(), index=df.index
    )

    # Drop the now unnecessary 'Index' column
    df.drop(columns=["Index"], inplace=True)

    return df


def escape_formula(text: str):
    """
    Prevent a string from being interpreted as a formula in Excel.
    Strings starting with '=', '-', or '+' are prefixed with an apostrophe.

    :param text: The string to be adjusted.
    :return: The adjusted string.
    """
    return "'" + text if text.startswith(("=", "-", "+")) else text


def run_GSA_OLS(methods: list, export_path: Path):
    """
    Runs OLS regression for specified methods and writes summaries to an Excel file.

    :param methods: Methods corresponding to dataset columns.
    :param export_path: Path to the directory where the Excel file will be saved.
    """
    try:
        book = load_workbook(export_path)
    except FileNotFoundError:
        book = Workbook()
        book.save(export_path)
        book = load_workbook(export_path)

    data = pd.read_excel(export_path, sheet_name="Sheet1")

    if "OLS" in book.sheetnames:
        ws = book["OLS"]
        book.remove(ws)

    ws = book.create_sheet("OLS")

    X_base = data.drop(columns=["Iteration", "Year"] + methods)
    X_base = sm.add_constant(X_base)
    corr_matrix = X_base.corr().abs()
    upper_triangle = corr_matrix.where(
        np.triu(np.ones(corr_matrix.shape), k=1).astype(bool)
    )
    high_correlation = [
        column
        for column in upper_triangle.columns
        if any(upper_triangle[column] > 0.95)
    ]

    if high_correlation:
        print(f"OLS: High multicollinearity detected in columns: {high_correlation}")
        X_base = X_base.drop(columns=high_correlation)

    results = []
    for method in methods:
        if method not in data.columns:
            print(f"Data for {method} not found in the file.")
            continue

        Y = data[method]
        X = X_base.copy()

        try:
            model_results = sm.OLS(Y, X).fit()
            summary = model_results.summary().as_text()
            summary_lines = summary.split("\n")

            results.append([f"OLS Summary for {method}"])
            for line in summary_lines:
                line = escape_formula(line)
                columns = re.split(r"\s{2,}", line)
                results.append(columns)
            results.append([])
        except Exception as e:
            print(f"Error running OLS for method {method}: {e}")

    for result in results:
        ws.append(result)

    book.save(export_path)


def run_GSA_delta(
    total_impacts: pd.DataFrame,
    uncertainty_values: pd.DataFrame,
    technology_shares: pd.DataFrame,
) -> pd.DataFrame:
    """
    Runs Delta Moment-Independent Measure analysis for specified methods and writes summaries to an Excel file.

    :param total_impacts: DataFrame with total impacts for each method.
    :param uncertainty_values: DataFrame with uncertainty values.
    :param technology_shares: DataFrame with technology shares.
    :return: DataFrame with Delta Moment-Independent Measure analysis results.
    """

    # merge uncertainty_values and technology_shares
    # based on "iteration" and "region" columns

    if len(technology_shares) > 0:
        df_parameters = uncertainty_values.merge(
            technology_shares, on=["iteration", "region"]
        )
    else:
        df_parameters = uncertainty_values

    parameters = [
        param for param in df_parameters.columns if param not in ["iteration", "region"]
    ]

    problem = {
        "num_vars": len(parameters),
        "names": parameters,
        "bounds": [
            [df_parameters[param].min(), df_parameters[param].max()]
            for param in parameters
        ],
    }

    methods = [m for m in total_impacts.columns if m not in ["iteration", "region"]]

    results = []

    for method in methods:
        param_values = df_parameters[parameters].values

        # total impacts for the method
        Y = total_impacts[method].values

        delta_results = delta.analyze(problem=problem, X=param_values, Y=Y)

        for i, param in enumerate(parameters):
            results.append(
                [
                    method,
                    param,
                    delta_results["delta"][i],
                    delta_results["delta_conf"][i],
                    delta_results["S1"][i],
                    delta_results["S1_conf"][i],
                ]
            )

    return pd.DataFrame(
        results,
        columns=["LCIA method", "Parameter", "Delta", "Delta Conf", "S1", "S1 Conf"],
    )


def log_mc_parameters_to_excel(
    model: str,
    scenario: str,
    year: int,
    methods: list,
    result: dict,
    uncertainty_parameters: dict,
    uncertainty_values: dict,
    tehnosphere_indices: dict,
    iteration_results: dict,
    shares: dict = None,
):
    export_path = STATS_DIR / f"{model}_{scenario}_{year}.xlsx"

    # create Excel workbook using openpyxl
    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:

        df_sum_impacts = pd.DataFrame()
        df_uncertainty_values = pd.DataFrame()
        df_technology_shares = pd.DataFrame()
        writer.book.create_sheet("Indices mapping")
        writer.book.create_sheet("Monte Carlo values")
        writer.book.create_sheet("Technology shares")
        writer.book.create_sheet("Total impacts")

        for region, data in result.items():

            total_impacts = np.sum(iteration_results[region], axis=(0, 2, 3))

            df_sum_impacts = pd.concat(
                [
                    df_sum_impacts,
                    log_results(
                        total_impacts=total_impacts,
                        methods=methods,
                        region=region,
                    ),
                ]
            )

            uncertainty_indices = uncertainty_parameters[region]
            uncertainty_vals = uncertainty_values[region]

            df_uncertainty_values = pd.concat(
                [
                    df_uncertainty_values,
                    log_uncertainty_values(
                        region=region,
                        uncertainty_indices=uncertainty_indices,
                        uncertainty_values=uncertainty_vals,
                    ),
                ],
            )

            if shares:
                sub_shares = {}
                for k, v in shares.items():
                    for x, y in v.items():
                        if x == year:
                            for z, w in y.items():
                                sub_shares[f"{k} - {z}"] = w

                df_technology_shares = pd.concat(
                    [
                        df_technology_shares,
                        log_subshares(
                            shares=sub_shares,
                            region=region,
                        ),
                    ],
                )

        indices = tehnosphere_indices[region]

        df_technosphere_indices = create_mapping_sheet(indices=indices)

        df_sum_impacts.to_excel(writer, sheet_name="Total impacts", index=False)
        df_uncertainty_values.to_excel(
            writer, sheet_name="Monte Carlo values", index=False
        )
        df_technology_shares.to_excel(
            writer, sheet_name="Technology shares", index=False
        )
        df_technosphere_indices.to_excel(
            writer, sheet_name="Indices mapping", index=False
        )

        print(f"Monte Carlo parameters added to: {export_path.resolve()}")


def run_gsa(directory: [str, None] = STATS_DIR, method: str = "delta") -> None:
    """
    Run a global sensitivity analysis (GSA) on the LCA results.
    Updates Excel files with the GSA results.
    :param method: str. The method used for the GSA. Default is 'delta'. Only 'delta' is supported at the moment.
    :param directory: str. The directory where the Excel files are stored. Default is 'stats'.
    :return: None.
    """
    if method != "delta":
        raise ValueError(f"Method {method} is not supported.")

    # iterate through the Excel files in the directory

    for file in Path(directory).rglob("*.xlsx"):
        # load content of "Monte Carlo values" sheet into a pandas DataFrame
        df_mc_vals = pd.read_excel(
            file, sheet_name="Monte Carlo values"
        )

        # load content of "Technology shares" sheet into a pandas DataFrame
        # if it exists

        try:
            df_technology_shares = pd.read_excel(
                file,
                sheet_name="Technology shares",
            )
        except:
            df_technology_shares = None

        # load content of "Total impacts" sheet into a pandas DataFrame
        df_sum_impacts = pd.read_excel(
            file, sheet_name="Total impacts"
        )

        # open Excel workbook
        with pd.ExcelWriter(
            file, engine="openpyxl", mode="a"
        ) as writer:

            df_GSA_results = run_GSA_delta(
                total_impacts=df_sum_impacts,
                uncertainty_values=df_mc_vals,
                technology_shares=df_technology_shares,
            )

            df_GSA_results.to_excel(
                writer, sheet_name=f"GSA {method.capitalize()}", index=False
            )

        print(f"GSA results added to: {file.resolve()}")
